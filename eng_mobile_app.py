import pandas as pd
import re
from collections import Counter
import os
import openpyxl
# 💡 [버그 수정 완료] 최신 openpyxl 버전에 맞게 Rich Text 임포트 경로 수정
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont 

def process_patterns(input_file, output_file, sheet_name='전체통합', target_col='영어'):
    """
    엑셀 파일의 영어 문장들을 분석하여 
    HTML 태그가 아닌 '실제 엑셀 굵은 글씨 서식(Rich Text)'을 적용합니다.
    """
    print(f"🔍 [{input_file}] 파일 분석을 시작합니다...")
    
    # 1. 엑셀 워크북 및 시트 로드 (openpyxl 사용)
    try:
        wb = openpyxl.load_workbook(input_file)
        if sheet_name not in wb.sheetnames:
            print(f"❌ 오류: '{sheet_name}' 시트를 찾을 수 없습니다.")
            return
        ws = wb[sheet_name]
    except Exception as e:
        print(f"❌ 엑셀 파일을 읽는 중 오류가 발생했습니다: {e}")
        return

    # 2. 타겟 컬럼 인덱스 찾기
    col_idx = None
    for cell in ws[1]: # 첫 번째 행을 헤더로 간주
        if cell.value == target_col:
            col_idx = cell.column
            break
            
    if not col_idx:
        print(f"❌ 오류: '{target_col}' 열을 찾을 수 없습니다.")
        return

    # 3. 데이터 읽기 및 문장 추출
    sentences = []
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=col_idx).value
        sentences.append(str(val) if val is not None else "")
    
    # 4. 문장 시작 부분의 패턴(n-gram) 추출
    patterns = []
    for text in sentences:
        clean_text = re.sub(r"[^\w\s']", ' ', text.lower()).strip()
        words = clean_text.split()
        
        for n in range(2, 6):
            if len(words) > n: # 문장 전체 길이보다 짧은 패턴만 추출
                ngram = " ".join(words[:n])
                patterns.append(ngram)
                
    # 5. 빈도수 계산 및 필터링
    counter = Counter(patterns)
    frequent_patterns = []
    
    for pat, count in counter.items():
        pat_len = len(pat.split())
        # 5번 이상 등장하는 의미 있는 패턴 모두 추출
        if pat_len == 2 and count >= 5:       
            frequent_patterns.append(pat)
        elif pat_len == 3 and count >= 5:     
            frequent_patterns.append(pat)
        elif pat_len >= 4 and count >= 5:     
            frequent_patterns.append(pat)
    
    # 긴 패턴부터 우선 매칭하기 위해 내림차순 정렬
    frequent_patterns.sort(key=lambda x: len(x.split()), reverse=True)
    
    print(f"✅ 스마트 필터링 결과, 총 {len(frequent_patterns)}개의 정제된 핵심 패턴을 찾았습니다.")

    # 6. 실제 엑셀 굵은 글씨 서식 (Rich Text) 적용
    print("⏳ 엑셀 파일에 '실제 굵은 글씨' 서식을 직접 입히고 있습니다. 잠시만 기다려주세요...")
    
    # 💡 굵은 글씨용 InlineFont 객체 생성
    bold_font = InlineFont(b=True) 

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        text_str = str(cell.value) if cell.value is not None else ""
        
        if not text_str.strip():
            continue
            
        clean_text = re.sub(r"[^\w\s']", ' ', text_str.lower()).strip()
        words_clean = clean_text.split()
        
        for pat in frequent_patterns:
            pat_words = pat.split()
            pat_len = len(pat_words)
            
            # 패턴과 일치하는 경우
            if len(words_clean) > pat_len and " ".join(words_clean[:pat_len]) == pat:
                word_pattern = r"[\w']+"         
                non_word_pattern = r"[^\w']*"    
                
                regex_str = r'^(' + non_word_pattern
                for _ in range(pat_len):
                    regex_str += word_pattern + non_word_pattern
                regex_str += r')'
                
                match = re.search(regex_str, text_str)
                if match:
                    matched_part = match.group(1)
                    m_clean = re.match(r'^(.*?[\w\'])([^\w\']*)$', matched_part)
                    
                    rich_text_elements = []
                    
                    if m_clean:
                        actual_text = m_clean.group(1) # 굵게 만들 순수 텍스트
                        trailing_chars = m_clean.group(2) # 뒤에 붙은 기호/공백
                        rest_of_text = text_str[len(matched_part):] # 나머지 텍스트
                        
                        # 태그가 아닌 실제 엑셀 서식 객체(TextBlock) 추가
                        rich_text_elements.append(TextBlock(font=bold_font, text=actual_text))
                        remaining_text = trailing_chars + rest_of_text
                        if remaining_text:
                            rich_text_elements.append(remaining_text)
                            
                    else:
                        rest_of_text = text_str[len(matched_part):]
                        rich_text_elements.append(TextBlock(font=bold_font, text=matched_part))
                        if rest_of_text:
                            rich_text_elements.append(rest_of_text)
                            
                    # 셀의 값을 Rich Text 객체로 덮어쓰기
                    cell.value = CellRichText(rich_text_elements)
                    break # 하나의 셀에는 하나의 가장 긴 패턴만 적용하고 다음 행으로 이동

    # 7. 새 엑셀 파일로 저장
    wb.save(output_file)
    print(f"🎉 작업 완료! 진짜 굵은 글씨가 적용된 [{output_file}] 파일이 생성되었습니다.")

if __name__ == "__main__":
    # =============== [설정 영역] ===============
    INPUT_EXCEL = "영어회화_통합본2.xlsx"  
    OUTPUT_EXCEL = "영어회화_통합본2_진짜굵은글씨.xlsx" 
    
    TARGET_SHEET = "전체통합" 
    TARGET_COLUMN = "영어"    
    # ===========================================
    
    if os.path.exists(INPUT_EXCEL):
        process_patterns(INPUT_EXCEL, OUTPUT_EXCEL, TARGET_SHEET, TARGET_COLUMN)
    else:
        print(f"⚠️ 현재 폴더에 '{INPUT_EXCEL}' 파일이 없습니다. 파일명과 위치를 다시 확인해 주세요.")
